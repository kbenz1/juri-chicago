import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

"""
Liest aus 'umfrageergebnisse.xlsx' die Worksheet-Daten per Pandas ein,
berechnet:
  - Durchschnitt New Yama-Cola,
  - Durchschnitt Yama-Cola Classic,
  - Differenz der beiden Durchschnitte,
  - Fehlerquote basierend auf 'Stichprobengrösse'.
Erstellt dann mit openpyxl in der Workbook-Kopie ein neues Worksheet
namens 'Auswertung' und schreibt die Ergebnisse als Key-Value-Tabelle
ohne Header. Passt zusätzlich die Spaltenbreite so an, dass
der Inhalt komplett sichtbar ist. Speichert das Ganze als 'auswertung.xlsx'.
"""


datei_einlesen = "umfrageergebnisse.xlsx"
# --- 1) Worksheet "Umfrageergebnisse" mit pandas lesen ---
df_umfrage = pd.read_excel(datei_einlesen, sheet_name="Umfrageergebnisse")

# Anzahl Testpersonen (Zeilen in 'Testperson')
testperson_count = df_umfrage["Testperson"].count()
print("Stichprobengrösse:", testperson_count)

# Durchschnitte berechnen
avg_new_yama = df_umfrage["Bewertung New Yama-Cola"].mean()
avg_classic = df_umfrage["Bewertung Yama-Cola Classic"].mean()
print("Durchschnitt New Yama-Cola:", avg_new_yama)
print("Durchschnitt alte Yama-Cola:", avg_classic)

# Differenz berechnen
differenz = avg_new_yama - avg_classic
print("Differenz:", differenz)

# --- 2) Worksheet "Fehlerquoten" lesen, Fehlerquote ermitteln ---
df_fehler = pd.read_excel(datei_einlesen, sheet_name="Fehlerquoten")

# Suche in 'Stichprobengrösse' nach dem Wert testperson_count
zeile = df_fehler.loc[df_fehler["Stichprobengrösse"] == testperson_count]

if not zeile.empty:
    fehlerquote_wert = zeile["Fehlerquote"].iloc[0]
else:
    fehlerquote_wert = 0

# --- 3) Dictionary mit allen Auswertungsergebnissen ---
ergebnisse = {
    "New Yama-Cola durchschnittl. Bewertung": avg_new_yama,
    "Alte Yama-Cola durchschnittl. Bewertung": avg_classic,
    "Differenz": differenz,
    "Fehlerquote": fehlerquote_wert
}

# --- 4) Kopie anlegen ---
oldwb = openpyxl.load_workbook(datei_einlesen)
datei_speichern = "auswertung.xlsx"
oldwb.save(datei_speichern)
oldwb.close()

wb = openpyxl.load_workbook(datei_speichern)
ws_auswertung = wb.create_sheet("Auswertung")

# Dictionary in zwei Spalten (Key / Value) eintragen, ohne Header
# und gleichzeitig Längen der Inhalte erfassen, um später Spaltenbreite anzupassen
row_idx = 1

# Wir merken uns die maximale Zeichenlänge in den Spalten 1 und 2
max_length_col1 = 0
max_length_col2 = 0

for key, value in ergebnisse.items():
    # Key und Value in die Zellen schreiben
    ws_auswertung.cell(row=row_idx, column=1, value=key)
    ws_auswertung.cell(row=row_idx, column=2, value=value)
    row_idx += 1
    # Zeichenzahl bestimmen, um die spätere Spaltenbreite anzupassen
    len_col1 = len(str(key))
    len_col2 = len(str(value))
    
    if len_col1 > max_length_col1:
        max_length_col1 = len_col1
    if len_col2 > max_length_col2:
        max_length_col2 = len_col2

# Spaltenbreite in Excel etwa an die Zeichenlänge anpassen
# (kleine Zugabe zur Breite, damit es nicht zu knapp ist)
ws_auswertung.column_dimensions[get_column_letter(1)].width = max_length_col1 + 2
ws_auswertung.column_dimensions[get_column_letter(2)].width = max_length_col2 + 2

# --- 5) Alles in eine neue Datei speichern und Workbook schließen ---
wb.save(datei_speichern)
wb.close()



