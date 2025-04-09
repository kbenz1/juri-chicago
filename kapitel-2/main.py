import pandas as pd
import openpyxl

def auswertung(datei_einlesen: str, datei_speichern: str):
    """
    Liest aus 'umfrageergebnisse.xlsx' die Worksheet-Daten per Pandas ein,
    berechnet:
      - Durchschnitt New Yama-Cola,
      - Durchschnitt Yama-Cola Classic,
      - Differenz der beiden Durchschnitte,
      - Fehlerquote basierend auf 'Stichprobengrösse'.
    Erstellt dann mit openpyxl in der Workbook-Kopie ein neues Worksheet
    namens 'Auswertung' und schreibt die Ergebnisse als Key-Value-Tabelle
    ohne Header. Speichert das Ganze als 'auswertung.xlsx'.
    """
    
    # --- 1) Worksheet "Umfrageergebnisse" mit pandas lesen ---
    df_umfrage = pd.read_excel(datei_einlesen, sheet_name="Umfrageergebnisse")
    
    # Entferne Zeilen ohne Testperson-Eintrag
    df_umfrage = df_umfrage.dropna(subset=["Testperson"])
    
    # Anzahl Testpersonen (Zeilen in 'Testperson')
    testperson_count = df_umfrage["Testperson"].count()
    
    # Durchschnitte berechnen
    avg_new_yama = df_umfrage["Bewertung New Yama-Cola"].mean()
    avg_classic = df_umfrage["Bewertung Yama-Cola Classic"].mean()
    
    differenz = avg_new_yama - avg_classic
    
    # --- 2) Worksheet "Fehlerquoten" lesen, Fehlerquote ermitteln ---
    df_fehler = pd.read_excel(datei_einlesen, sheet_name="Fehlerquoten")
    
    # Suche in 'Stichprobengrösse' nach dem Wert testperson_count
    zeile = df_fehler.loc[df_fehler["Stichprobengrösse"] == testperson_count]
    
    if not zeile.empty:
        fehlerquote_wert = zeile["Fehlerquote"].iloc[0]
    else:
        fehlerquote_wert = 0
    
    # --- 3) Dictionary mit allen Auswertungsergebnissen ---
    ergebnisse = {
        "New Yama-Cola durchschnittl. Bewertung": avg_new_yama,
        "Alte Yama-Cola durchschnittl. Bewertung": avg_classic,
        "Differenz": differenz,
        "Fehlerquote": fehlerquote_wert
    }
    
    # --- 4) Neues Worksheet "Auswertung" in der Originaldatei anlegen ---
    #     Danach als neues Excelfile 'auswertung.xlsx' speichern
    wb = openpyxl.load_workbook(datei_einlesen)
    
    # Falls das Sheet "Auswertung" evtl. schon existiert, vorab entfernen
    if "Auswertung" in wb.sheetnames:
        del wb["Auswertung"]
    
    ws_auswertung = wb.create_sheet("Auswertung")
    
    # Dictionary in zwei Spalten (Key / Value) eintragen, ohne Header
    row_idx = 1
    for key, value in ergebnisse.items():
        ws_auswertung.cell(row=row_idx, column=1, value=key)
        ws_auswertung.cell(row=row_idx, column=2, value=value)
        row_idx += 1
    
    # --- 5) Alles in eine neue Datei speichern und Workbook schließen ---
    wb.save(datei_speichern)
    wb.close()

if __name__ == "__main__":
    # Beispielaufruf: 
    auswertung("umfrageergebnisse.xlsx", "auswertung.xlsx")
